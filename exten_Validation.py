import openpyxl
import re
from collections import defaultdict

def rgb_to_hex(rgb):
    return ''.join(f'{v:02X}' for v in rgb)

def get_fill_color(cell):
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        return fill.start_color.rgb[-6:]  # Take last 6 chars as RGB hex without alpha
    return None

def infer_pattern(values):
    """Infer pattern type from a list of non-null string values."""
    numeric = True
    number_with_unit = True
    alphabetic = True
    multi_value = False

    for v in values:
        if not isinstance(v, str):
            v = str(v)
        v_strip = v.strip()
        # Numeric check
        if numeric:
            if not re.fullmatch(r'\d+(\.\d+)?', v_strip):
                numeric = False
        # Number with unit check (e.g. "5 hp")
        if number_with_unit:
            if not re.fullmatch(r'\d+(\.\d+)?\s*\w+', v_strip):
                number_with_unit = False
        # Alphabetic check (letters, spaces)
        if alphabetic:
            if not re.fullmatch(r'[A-Za-z\s]+', v_strip):
                alphabetic = False
        # Multi-value check (comma separated)
        if ',' in v_strip:
            multi_value = True

    if multi_value:
        return 'multi_value_text'
    if numeric:
        return 'numeric'
    if number_with_unit:
        return 'number_with_unit'
    if alphabetic:
        return 'text'
    return 'mixed'

def fix_quotes(val):
    if not isinstance(val, str):
        return val
    val = val.strip()
    # Remove leading and trailing single/double quotes repeatedly
    while (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
        val = val[1:-1].strip()
    return val

def clean_commas(val):
    if not isinstance(val, str):
        return val
    # Remove leading/trailing commas and spaces around delimiters
    val = val.strip()
    val = re.sub(r'^[,]+\s*', '', val)
    val = re.sub(r'\s*[,]+$', '', val)
    # Replace multiple commas or mixed delimiters with a single comma no space
    val = re.sub(r'[;|/]', ',', val)
    val = re.sub(r'\s*,\s*', ',', val)
    return val

def standardize_case(val, allowed_values):
    # val string, allowed_values is list/set including cases
    if not isinstance(val, str):
        return val
    clean_val = val.strip()
    if clean_val in allowed_values:
        return clean_val
    # Match ignoring case
    matches = [v for v in allowed_values if v.lower() == clean_val.lower()]
    if not matches:
        return val  # no match, no change
    # Prefer uppercase if any match fully uppercase
    for m in matches:
        if m.isupper():
            return m
    # else return first matched (case-insensitive)
    return matches[0]

def cell_value_matches_pattern(val, pattern):
    if val is None:
        return True
    if pattern == 'numeric':
        # Accept integer or float string
        if isinstance(val, (int, float)):
            return True
        if isinstance(val, str) and re.fullmatch(r'\d+(\.\d+)?', val.strip()):
            return True
        return False
    if pattern == 'number_with_unit':
        if isinstance(val, str) and re.fullmatch(r'\d+(\.\d+)?\s*\w+', val.strip()):
            return True
        return False
    if pattern == 'text':
        if isinstance(val, str) and re.fullmatch(r'[A-Za-z\s]+', val.strip()):
            return True
        return False
    if pattern == 'multi_value_text':
        # Check all sub-values satisfy text pattern
        if not isinstance(val, str):
            return False
        parts = [p.strip() for p in val.split(',')]
        for p in parts:
            if not re.fullmatch(r'[A-Za-z\s]+', p):
                return False
        return True
    # For mixed or unknown pattern, accept all
    return True

def extract_price_range(sheet1, price_col_idx):
    # Scan Sheet1 for min and max price numeric values ignoring empty and non-numeric
    prices = []
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        val = row[price_col_idx]
        if val is None:
            continue
        s = str(val).strip().replace('$','').replace(',','')
        try:
            f = float(s)
            prices.append(f)
        except:
            pass
    if not prices:
        return None, None
    return min(prices), max(prices)

def run_validation_all(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb['Sheet1']
    sheet2 = wb['Data']

    headers_sheet1 = [cell.value for cell in sheet1[1]]
    headers_sheet2 = [cell.value for cell in sheet2[1]]

    # Identify green color hex to filter columns in data sheet (header row only)
    green_hex = '00B050'

    # Find green highlighted columns in data sheet header
    green_col_indices = []
    for idx, cell in enumerate(sheet2[1], start=1):
        fill_color = cell.fill.start_color
        # Checking RGB against our green.
        # Depending on Excel version, fill_color.rgb or fill_color.indexed may be set
        if fill_color.type == 'rgb' and fill_color.rgb[-6:] == green_hex:
            green_col_indices.append(idx)
        elif fill_color.type == 'indexed':
            # Fallback, approximate check for green (Excel indexed green is 10)
            if fill_color.indexed == 10:
                green_col_indices.append(idx)

    # Map green column indices to their headers
    green_headers = [sheet2.cell(row=1, column=i).value for i in green_col_indices]

    # Intersection with Sheet1 columns
    common_columns = list(set(green_headers).intersection(headers_sheet1))

    # Build allowed values dictionary for columns - including subvalues for groups
    allowed_values = {}
    for col in common_columns:
        idx1 = headers_sheet1.index(col)
        vals = set()
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            val = row[idx1]
            if val is not None:
                str_val = str(val).strip()
                vals.add(str_val)
                # Also add comma-separated elements as allowed
                if ',' in str_val:
                    for subval in str_val.split(','):
                        vals.add(subval.strip())
        allowed_values[col] = vals

    # Infer column patterns from Sheet1 values only for common columns
    column_patterns = {}
    for col in common_columns:
        idx1 = headers_sheet1.index(col)
        vals = [str(r[idx1]).strip() for r in sheet1.iter_rows(min_row=2, values_only=True) if r[idx1] is not None]
        column_patterns[col] = infer_pattern(vals) if vals else 'unknown'

    # Find price column index and range if present
    price_col_idx_sheet1 = None
    price_col_idx_data = None
    min_price, max_price = None, None
    for col in common_columns:
        if col.lower() == 'price':
            price_col_idx_sheet1 = headers_sheet1.index(col)
            price_col_idx_data = sheet2[1].index(col)+1
            min_price, max_price = extract_price_range(sheet1, price_col_idx_sheet1)
            break

    # Add Comments column if missing or get index
    if 'Comments' not in headers_sheet2:
        comments_col_idx = len(headers_sheet2) + 1
        sheet2.cell(row=1, column=comments_col_idx, value='Comments')
    else:
        comments_col_idx = headers_sheet2.index('Comments') + 1

    error_counters = defaultdict(int)
    total_cells_checked = 0

    # For each row and valid column do validation
    for row_num, row in enumerate(sheet2.iter_rows(min_row=2), 2):
        row_errors = []
        for col in common_columns:
            col_idx_data = headers_sheet2.index(col)
            cell = row[col_idx_data]
            val = cell.value

            # Clean quotes and trim whitespace
            if isinstance(val, str):
                old_val = val
                val = fix_quotes(val)
                val = val.strip()
                val = clean_commas(val)
                # Replace non-standard delimiters ; | / with commas and remove spaces around commas
                val = re.sub(r'[;|/]', ',', val)
                val = re.sub(r'\s*,\s*', ',', val)
                if val != old_val:
                    cell.value = val

            total_cells_checked += 1

            # Case correction: map to allowed value with case preference like uppercase if both
            if isinstance(val, str) and col in allowed_values:
                mapped_val = standardize_case(val, allowed_values[col])
                if mapped_val != val:
                    cell.value = mapped_val
                    val = mapped_val
                    row_errors.append(f'{col}: Case corrected')

            # Check pattern match
            if not cell_value_matches_pattern(val, column_patterns[col]):
                row_errors.append(f'{col}: Pattern mismatch')
                error_counters['pattern_mismatch'] +=1

            # Check duplicates in multi-value cells
            if isinstance(val, str) and ',' in val:
                parts = [p.strip() for p in val.split(',') if p.strip()]
                if len(parts) != len(set(parts)):
                    row_errors.append(f'{col}: Duplicates values in cell')
                    error_counters['duplicates'] += 1

            # Check allowed values
            if isinstance(val, str) and col in allowed_values:
                parts = [p.strip() for p in val.split(',') if p.strip()]
                for p in parts:
                    if p not in allowed_values[col]:
                        row_errors.append(f'{col}: Value "{p}" not allowed')
                        error_counters['invalid_value'] += 1
            elif val is not None and col in allowed_values and val not in allowed_values[col]:
                row_errors.append(f'{col}: Value "{val}" not allowed')
                error_counters['invalid_value'] += 1

            # Numeric range check on price column
            if price_col_idx_data and col == 'Price':
                # convert val to float if possible
                try:
                    num_val = float(str(val).replace('$','').replace(',','').strip())
                    if min_price is not None and num_val < min_price:
                        row_errors.append(f'{col}: Below min price {min_price}')
                        error_counters['price_range'] += 1
                    if max_price is not None and num_val > max_price:
                        row_errors.append(f'{col}: Above max price {max_price}')
                        error_counters['price_range'] += 1
                except:
                    row_errors.append(f'{col}: Price not a number')
                    error_counters['price_range'] += 1

            # Numeric text validations - no trailing .0, max two decimals
            if isinstance(val, str) and re.fullmatch(r'\d+(\.\d+)?', val):
                if val.endswith('.0'):
                    row_errors.append(f'{col}: Numeric value ends with .0')
                    error_counters['numeric_format'] += 1
                if '.' in val:
                    dec = val.split('.')[1]
                    if len(dec) > 2:
                        row_errors.append(f'{col}: Numeric value has more than two decimals')
                        error_counters['numeric_format'] += 1

            # Remove special chars at start/end - apply again if needed
            if isinstance(val, str):
                if re.match(r'^[^A-Za-z0-9]+', val) or re.match(r'[^A-Za-z0-9]+$', val):
                    cleaned = val.strip(' !@#$%^&*()_+-=[]{};:\'",.<>?/|\\')
                    if cleaned != val:
                        cell.value = cleaned
                        val = cleaned
                        row_errors.append(f'{col}: Trimmed special chars')

            # Detect Excel formulas
            if isinstance(val, str) and val.startswith('='):
                row_errors.append(f'{col}: Contains formula')
                error_counters['formula'] += 1

        # Compile message in comments
        if row_errors:
            sheet2.cell(row=row_num, column=comments_col_idx).value = ', '.join(row_errors)
            for err in set(row_errors):
                error_counters[err.split(':')[0].lower()] += 1

    # Report error percentages on summary sheet
    summary_sheet = wb.create_sheet('Validation_Summary')
    summary_sheet.append(['Error Type', 'Count', 'Percentage'])
    total = total_cells_checked if total_cells_checked > 0 else 1
    for error, count in error_counters.items():
        pct = round((count / total) * 100, 2)
        summary_sheet.append([error, count, pct])

    wb.save(output_path)

if __name__ == '__main__':
    input_file = 'Industrial-Automation-and-Controls_AC-Motors_PDW_[by_Sarang-P]_1762946783_6b35238e.xlsx'
    output_file = 'validated_full_report.xlsx'
    run_validation_all(input_file, output_file)
    print(f'Validation completed and saved in {output_file}')
