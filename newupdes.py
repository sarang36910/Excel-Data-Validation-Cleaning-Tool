import openpyxl
import re
import time
from collections import defaultdict

def get_actual_data_limits(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, ws.max_column + 1)):
            max_row = r
            break
    for c in range(ws.max_column, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for r in range(1, ws.max_row + 1)):
            max_col = c
            break
    return max_row, max_col

def fix_quotes(val):
    if not isinstance(val, str):
        return val, False
    val_orig = val
    val = val.strip()
    while (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
        val = val[1:-1].strip()
    return val, val != val_orig

def clean_commas_and_empty(val):
    if not isinstance(val, str):
        return val, False
    val_orig = val
    val = val.strip()
    parts = [p.strip() for p in val.split(',') if p.strip() != '']
    val_cleaned = ','.join(parts)
    changed = val_cleaned != val_orig
    return val_cleaned, changed

def standardize_case(val, allowed_values):
    if not isinstance(val, str):
        return val, False
    clean_val = val.strip()
    if clean_val in allowed_values:
        return clean_val, clean_val != val
    matches = [v for v in allowed_values if v.lower() == clean_val.lower()]
    if not matches:
        return val, False
    for m in matches:
        if m.isupper():
            return m, m != val
    return matches[0], matches[0] != val

def extract_extensions(values):
    exts = {}
    pattern = re.compile(r'(\d+(?:\.\d+)?)\s*(\w+)', re.I)
    nums_per_ext = defaultdict(list)
    for v in values:
        m = pattern.fullmatch(v)
        if m:
            num = float(m.group(1))
            ext = m.group(2).lower()
            nums_per_ext[ext].append(num)
    for ext, nums in nums_per_ext.items():
        exts[ext] = (min(nums), max(nums))
    return exts

def parse_number_and_extension(val):
    pattern = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*(\w*)\s*$', re.I)
    m = pattern.match(val)
    if m:
        number = m.group(1)
        extension = m.group(2)
        return float(number), extension.lower()
    return None, None

def run_validation_all(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb['Sheet1']
    sheet2 = wb['Data']

    max_row1, max_col1 = get_actual_data_limits(sheet1)
    max_row2, max_col2 = get_actual_data_limits(sheet2)

    headers_sheet1 = [cell.value for cell in sheet1[1][:max_col1]]
    headers_sheet2 = [cell.value for cell in sheet2[1][:max_col2]]

    green_hex = '00B050'
    green_col_indices = []
    for idx, cell in enumerate(sheet2[1][:max_col2], start=1):
        fill_color = cell.fill.start_color
        if fill_color.type == 'rgb' and fill_color.rgb and fill_color.rgb[-6:].upper() == green_hex:
            green_col_indices.append(idx)
        elif fill_color.type == 'indexed':
            if fill_color.indexed == 10:
                green_col_indices.append(idx)
    green_headers = [sheet2.cell(row=1, column=i).value for i in green_col_indices]

    common_columns = list(set(green_headers).intersection(headers_sheet1))

    allowed_values = {}
    numeric_extension_info = {}
    idx_map_sheet1 = {}
    for col in common_columns:
        idx1 = headers_sheet1.index(col)
        idx_map_sheet1[col] = idx1
        vals = set()
        vals_for_ext = []
        for row in sheet1.iter_rows(min_row=2, max_row=max_row1, max_col=max_col1, values_only=True):
            val = row[idx1]
            if val is not None:
                str_val = str(val).strip()
                if ',' in str_val:
                    continue
                vals.add(str_val)
                vals_for_ext.append(str_val)
        allowed_values[col] = vals

        exts = extract_extensions(vals_for_ext)
        if exts:
            numeric_extension_info[col] = exts

    # Add Comments and Updates columns if missing
    if 'Comments' not in headers_sheet2:
        comments_col_idx = max_col2 + 1
        sheet2.cell(row=1, column=comments_col_idx, value='Comments')
    else:
        comments_col_idx = headers_sheet2.index('Comments') + 1

    if 'Updates Here' not in headers_sheet2:
        updates_col_idx = max_col2 + 2 if 'Comments' not in headers_sheet2 else max_col2 + 1
        sheet2.cell(row=1, column=updates_col_idx, value='Updates Here')
    else:
        updates_col_idx = headers_sheet2.index('Updates Here') + 1

    error_counters = defaultdict(int)
    total_cells_checked = 0

    for row_num, row in enumerate(sheet2.iter_rows(min_row=2, max_row=max_row2, max_col=max_col2), start=2):
        row_errors = []
        row_updates = []
        for col in common_columns:
            col_idx_data = headers_sheet2.index(col)
            cell = row[col_idx_data]
            val = cell.value

            # Remove quotes, add comment on removal
            if isinstance(val, str):
                val_new, quotes_removed = fix_quotes(val)
                if quotes_removed:
                    row_updates.append(f'{col}: Removed quotes')
                    val = val_new
                    cell.value = val

            # Clean commas, remove empty entries before/after commas, comment if changed
            if isinstance(val, str):
                val_new, commas_fixed = clean_commas_and_empty(val)
                if commas_fixed:
                    row_updates.append(f'{col}: Removed empty values and fixed commas')
                    val = val_new
                    cell.value = val

            total_cells_checked += 1

            # Case correction for individual items in multi-value cells
            if isinstance(val, str) and col in allowed_values:
                parts = [p.strip() for p in val.split(',')]
                corrected_parts = []
                changed_case = False
                for part in parts:
                    corrected_part, changed = standardize_case(part, allowed_values[col])
                    corrected_parts.append(corrected_part)
                    if changed:
                        changed_case = True
                corrected_val = ','.join(corrected_parts)
                if changed_case:
                    cell.value = corrected_val
                    val = corrected_val
                    row_updates.append(f'{col}: Case corrected on values')

            # Numeric + extension special logic
            if col in numeric_extension_info:
                num_val, ext_val = parse_number_and_extension(str(val) if val is not None else '')
                exts_allowed = numeric_extension_info[col]
                if num_val is not None:
                    if ext_val == '' and len(exts_allowed) == 1:
                        ext_val = next(iter(exts_allowed))
                        val_new = f"{int(num_val) if num_val.is_integer() else num_val} {ext_val}"
                        cell.value = val_new
                        val = val_new
                        row_updates.append(f'{col}: Added missing extension "{ext_val}"')
                    elif ext_val != '' and ext_val not in exts_allowed:
                        row_errors.append(f'{col}: Extension "{ext_val}" not standard but accepted')
                    if ext_val in exts_allowed:
                        min_n, max_n = exts_allowed[ext_val]
                        if num_val < min_n or num_val > max_n:
                            row_errors.append(f'{col}: Numeric value {num_val} exceeds allowed range [{min_n}, {max_n}]')
                else:
                    if val not in allowed_values[col]:
                        row_errors.append(f'{col}: Value "{val}" not allowed')
            else:
                # Numeric only values: check range and provide warning, but don't validate presence in sheet1
                if val is not None:
                    try:
                        num_val = float(str(val).strip())
                        if col in numeric_extension_info and "" in numeric_extension_info[col]:
                            min_n, max_n = numeric_extension_info[col][""]
                            if num_val < min_n or num_val > max_n:
                                row_errors.append(f'{col}: Numeric value {num_val} exceeds allowed range [{min_n}, {max_n}]')
                        else:
                            row_errors.append(f'{col}: Numeric value without extension found')
                    except:
                        if isinstance(val, str):
                            values = [v.strip() for v in val.split(',')]
                            values = [v for v in values if v != '']
                            if len(values) != len(set(values)):
                                row_errors.append(f'{col}: Duplicated values in cell')
                                error_counters['duplicates'] += 1
                            for v in values:
                                if v not in allowed_values[col]:
                                    row_errors.append(f'{col}: Value "{v}" not allowed')
                                    error_counters['invalid_value'] += 1
                        else:
                            if str(val) not in allowed_values[col]:
                                row_errors.append(f'{col}: Value "{val}" not allowed')
                                error_counters['invalid_value'] += 1

            # Numeric .0 and decimal places check
            if isinstance(val, str) and re.fullmatch(r'\d+(\.\d+)?', val):
                if val.endswith('.0'):
                    row_errors.append(f'{col}: Numeric value ends with .0')
                    error_counters['numeric_format'] += 1
                if '.' in val:
                    dec = val.split('.')[1]
                    if len(dec) > 2:
                        row_errors.append(f'{col}: Numeric value has more than two decimals')
                        error_counters['numeric_format'] += 1

            # Trim special characters
            if isinstance(val, str):
                if re.match(r'^[^A-Za-z0-9]+', val) or re.match(r'[^A-Za-z0-9]+$', val):
                    cleaned = val.strip(' !@#$%^&*()_+-=[]{};:\'",.<>?/|\\')
                    if cleaned != val:
                        cell.value = cleaned
                        val = cleaned
                        row_updates.append(f'{col}: Trimmed special chars')

            # Detect formulas
            if isinstance(val, str) and val.startswith('='):
                row_errors.append(f'{col}: Contains formula')
                error_counters['formula'] += 1

        # Write Updates and Comments columns
        if row_updates:
            prev_updates = sheet2.cell(row=row_num, column=updates_col_idx).value or ""
            sheet2.cell(row=row_num, column=updates_col_idx).value = (prev_updates + ', ' if prev_updates else '') + ', '.join(row_updates)
        if row_errors:
            prev_comments = sheet2.cell(row=row_num, column=comments_col_idx).value or ""
            sheet2.cell(row=row_num, column=comments_col_idx).value = (prev_comments + ', ' if prev_comments else '') + ', '.join(row_errors)
            for err in set(row_errors):
                error_counters[err.split(':')[0].lower()] += 1

    summary_sheet = wb.create_sheet('Validation_Summary')
    summary_sheet.append(['Error Type', 'Count', 'Percentage'])
    total = total_cells_checked if total_cells_checked > 0 else 1
    for error, count in error_counters.items():
        pct = round((count / total) * 100, 2)
        summary_sheet.append([error, count, pct])

    wb.save(output_path)

if __name__ == '__main__':
    import time
    input_file = 'IAC_AC-Drives_reverse_PDW_(by_Steffy-Senson)_1763094814_14fc87e6_Allocation_file_Nov-14.xlsx'  # change file path
    output_file = 'validated_report_with_updates_and_comments_with_time.xlsx'
    start_time = time.time()
    run_validation_all(input_file, output_file)
    end_time = time.time()
    print(f'Validation completed in {end_time - start_time:.2f} seconds and saved to {output_file}')
