import openpyxl
import re
from collections import defaultdict

def get_actual_data_limits(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    # Adjust max_row to last row with any data
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, ws.max_column + 1)):
            max_row = r
            break
    # Adjust max_col to last column with any data
    for c in range(ws.max_column, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for r in range(1, ws.max_row + 1)):
            max_col = c
            break
    return max_row, max_col

def fix_quotes(val):
    if not isinstance(val, str):
        return val
    val = val.strip()
    while (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
        val = val[1:-1].strip()
    return val

def clean_commas(val):
    if not isinstance(val, str):
        return val
    val = val.strip()
    val = re.sub(r'^[,]+\s*', '', val)
    val = re.sub(r'\s*[,]+$', '', val)
    val = re.sub(r'[;|/]', ',', val)
    val = re.sub(r'\s*,\s*', ',', val)
    return val

def standardize_case(val, allowed_values):
    if not isinstance(val, str):
        return val
    clean_val = val.strip()
    if clean_val in allowed_values:
        return clean_val
    matches = [v for v in allowed_values if v.lower() == clean_val.lower()]
    if not matches:
        return val
    for m in matches:
        if m.isupper():
            return m
    return matches[0]

def extract_price_range(sheet1, price_col_idx, max_row, max_col):
    prices = []
    for row in sheet1.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
        val = row[price_col_idx]
        if val is None:
            continue
        s = str(val).strip().replace('$','').replace(',','')
        try:
            f = float(s)
            prices.append(f)
        except:
            pass
    if not prices:
        return None, None
    return min(prices), max(prices)

def run_validation_all(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb['Sheet1']
    sheet2 = wb['Data']

    max_row1, max_col1 = get_actual_data_limits(sheet1)
    max_row2, max_col2 = get_actual_data_limits(sheet2)

    headers_sheet1 = [cell.value for cell in sheet1[1][:max_col1]]
    headers_sheet2 = [cell.value for cell in sheet2[1][:max_col2]]

    green_hex = '00B050'
    green_col_indices = []
    for idx, cell in enumerate(sheet2[1][:max_col2], start=1):
        fill_color = cell.fill.start_color
        if fill_color.type == 'rgb' and fill_color.rgb and fill_color.rgb[-6:].upper() == green_hex:
            green_col_indices.append(idx)
        elif fill_color.type == 'indexed':
            if fill_color.indexed == 10:
                green_col_indices.append(idx)

    green_headers = [sheet2.cell(row=1, column=i).value for i in green_col_indices]

    common_columns = list(set(green_headers).intersection(headers_sheet1))

    allowed_values = {}
    idx_map_sheet1 = {}
    for col in common_columns:
        idx1 = headers_sheet1.index(col)
        idx_map_sheet1[col] = idx1
        vals = set()
        for row in sheet1.iter_rows(min_row=2, max_row=max_row1, max_col=max_col1, values_only=True):
            val = row[idx1]
            if val is not None:
                str_val = str(val).strip()
                vals.add(str_val)
                if ',' in str_val:
                    for subval in str_val.split(','):
                        vals.add(subval.strip())
        allowed_values[col] = vals

    price_col_idx_sheet1 = None
    price_col_idx_data = None
    min_price, max_price = None, None
    for col in common_columns:
        if col.lower() == 'price':
            price_col_idx_sheet1 = headers_sheet1.index(col)
            price_col_idx_data = headers_sheet2.index(col)
            min_price, max_price = extract_price_range(sheet1, price_col_idx_sheet1, max_row1, max_col1)
            break

    if 'Comments' not in headers_sheet2:
        comments_col_idx = max_col2 + 1
        sheet2.cell(row=1, column=comments_col_idx, value='Comments')
    else:
        comments_col_idx = headers_sheet2.index('Comments') + 1

    error_counters = defaultdict(int)
    total_cells_checked = 0

    for row_num, row in enumerate(sheet2.iter_rows(min_row=2, max_row=max_row2, max_col=max_col2), start=2):
        row_errors = []
        for col in common_columns:
            col_idx_data = headers_sheet2.index(col)
            cell = row[col_idx_data]
            val = cell.value

            if isinstance(val, str):
                old_val = val
                val = fix_quotes(val)
                val = val.strip()
                val = clean_commas(val)
                val = re.sub(r'[;|/]', ',', val)
                val = re.sub(r'\s*,\s*', ',', val)
                if val != old_val:
                    cell.value = val

            total_cells_checked += 1

            if isinstance(val, str) and col in allowed_values:
                mapped_val = standardize_case(val, allowed_values[col])
                if mapped_val != val:
                    cell.value = mapped_val
                    val = mapped_val
                    row_errors.append(f'{col}: Case corrected')

            if isinstance(val, str) and ',' in val:
                parts = [p.strip() for p in val.split(',') if p.strip()]
                if len(parts) != len(set(parts)):
                    row_errors.append(f'{col}: Duplicates values in cell')
                    error_counters['duplicates'] += 1

            if isinstance(val, str) and col in allowed_values:
                parts = [p.strip() for p in val.split(',') if p.strip()]
                for p in parts:
                    if p not in allowed_values[col]:
                        row_errors.append(f'{col}: Value "{p}" not allowed')
                        error_counters['invalid_value'] += 1
            elif val is not None and col in allowed_values and val not in allowed_values[col]:
                row_errors.append(f'{col}: Value "{val}" not allowed')
                error_counters['invalid_value'] += 1

            if price_col_idx_data is not None and col.lower() == 'price':
                try:
                    num_val = float(str(val).replace('$','').replace(',','').strip())
                    if min_price is not None and num_val < min_price:
                        row_errors.append(f'{col}: Below min price {min_price}')
                        error_counters['price_range'] += 1
                    if max_price is not None and num_val > max_price:
                        row_errors.append(f'{col}: Above max price {max_price}')
                        error_counters['price_range'] += 1
                except:
                    row_errors.append(f'{col}: Price not a number')
                    error_counters['price_range'] += 1

            if isinstance(val, str) and re.fullmatch(r'\d+(\.\d+)?', val):
                if val.endswith('.0'):
                    row_errors.append(f'{col}: Numeric value ends with .0')
                    error_counters['numeric_format'] += 1
                if '.' in val:
                    dec = val.split('.')[1]
                    if len(dec) > 2:
                        row_errors.append(f'{col}: Numeric value has more than two decimals')
                        error_counters['numeric_format'] += 1

            if isinstance(val, str):
                if re.match(r'^[^A-Za-z0-9]+', val) or re.match(r'[^A-Za-z0-9]+$', val):
                    cleaned = val.strip(' !@#$%^&*()_+-=[]{};:\'",.<>?/|\\')
                    if cleaned != val:
                        cell.value = cleaned
                        val = cleaned
                        row_errors.append(f'{col}: Trimmed special chars')

            if isinstance(val, str) and val.startswith('='):
                row_errors.append(f'{col}: Contains formula')
                error_counters['formula'] += 1

        if row_errors:
            sheet2.cell(row=row_num, column=comments_col_idx).value = ', '.join(row_errors)
            for err in set(row_errors):
                error_counters[err.split(':')[0].lower()] += 1

    summary_sheet = wb.create_sheet('Validation_Summary')
    summary_sheet.append(['Error Type', 'Count', 'Percentage'])
    total = total_cells_checked if total_cells_checked > 0 else 1
    for error, count in error_counters.items():
        pct = round((count / total) * 100, 2)
        summary_sheet.append([error, count, pct])

    wb.save(output_path)

if __name__ == '__main__':
    input_file = 'Industrial-Automation-and-Controls_AC-Motors_PDW_[by_Sarang-P]_1762946783_6b35238e.xlsx'  # Update if needed
    output_file = 'validated_full_report_optimized.xlsx'
    run_validation_all(input_file, output_file)
    print(f'Validation completed and saved in {output_file}')


# if __name__ == '__main__':
#     input_file = 'Industrial-Automation-and-Controls_AC-Motors_PDW_[by_Sarang-P]_1762946783_6b35238e.xlsx'
#     output_file = 'validated_full_report_no_pattern.xlsx'
#     run_validation_all(input_file, output_file)
#     print(f'Validation completed and saved in {output_file}')
