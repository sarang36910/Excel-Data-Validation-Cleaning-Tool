import openpyxl
import re

def run_validations(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    sheet1 = wb['Sheet1']
    sheet2 = wb['Data']

    headers_sheet1 = [cell.value for cell in sheet1[1]]
    headers_sheet2 = [cell.value for cell in sheet2[1]]

    # Find common columns
    common_columns = list(set(headers_sheet1).intersection(set(headers_sheet2)))

    # Create allowed values dictionary
    allowed_values = {}
    for col in common_columns:
        idx = headers_sheet1.index(col)
        vals = set()
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            val = row[idx]
            if val is not None:
                vals.add(val)
        allowed_values[col] = vals

    # Add Comments column if not present
    if 'Comments' not in headers_sheet2:
        comments_col_idx = len(headers_sheet2) + 1
        sheet2.cell(row=1, column=comments_col_idx, value='Comments')
    else:
        comments_col_idx = headers_sheet2.index('Comments') + 1

    # Units columns example (customize as needed)
    units_columns = ['hp']  # example column with units

    def append_issue(row_num, issue):
        comment_cell = sheet2.cell(row=row_num, column=comments_col_idx)
        existing = comment_cell.value
        if existing:
            comment_cell.value = existing + ', ' + issue
        else:
            comment_cell.value = issue

    for row_num, row in enumerate(sheet2.iter_rows(min_row=2), 2):
        row_issues = []

        for col_name in common_columns:
            col_idx_2 = headers_sheet2.index(col_name)
            cell = row[col_idx_2]
            val = cell.value
            allowed_vals = allowed_values[col_name]

            # Trim whitespace in text cells
            if isinstance(val, str):
                trimmed = val.strip()
                if trimmed != val:
                    val = trimmed
                    cell.value = trimmed
                    row_issues.append(f'{col_name}: Whitespace trimmed')

            # Convert trailing .0 floats to int, when possible
            if isinstance(val, float) and val.is_integer():
                cell.value = int(val)
                val = int(val)
                row_issues.append(f'{col_name}: Converted to integer')

            # Check exact allowed values for all columns with possibly multiple comma-separated values
            if val is not None:
                if isinstance(val, str):
                    # Split on comma for multi-value check
                    values = [v.strip() for v in val.split(',') if v.strip()]
                    # Check for duplicates in cell
                    if len(values) != len(set(values)):
                        row_issues.append(f'{col_name}: Duplicate value in cell')
                    # Check if all values are in allowed set
                    for v in values:
                        if v not in allowed_vals:
                            row_issues.append(f'{col_name}: Value \"{v}\" not allowed')
                else:
                    # For non-string types, simple match check
                    if val not in allowed_vals:
                        row_issues.append(f'{col_name}: Value \"{val}\" not allowed')

            # Unit column check
            if col_name in units_columns and isinstance(val, str):
                match = re.match(r'^(\d+(?:\.\d+)?)\s*(\w+)$', val)
                if not match:
                    row_issues.append(f'{col_name}: Missing or invalid unit')
                else:
                    number, unit = match.groups()
                    if number.endswith('.0'):
                        row_issues.append(f'{col_name}: Number has trailing .0')

            # Multi-value delimiter spacing fix
            if isinstance(val, str) and ',' in val:
                if ', ' in val or ' ,' in val:
                    fixed_val = val.replace(', ', ',').replace(' ,', ',')
                    cell.value = fixed_val
                    row_issues.append(f'{col_name}: Fixed spaces around commas')

            # Detect formulas
            if isinstance(val, str) and val.startswith('='):
                row_issues.append(f'{col_name}: Contains formula')

            # Replace non-standard delimiters
            if isinstance(val, str):
                new_val = val
                for delim in [';', '|', '/']:
                    if delim in new_val:
                        new_val = new_val.replace(delim, ',')
                new_val = re.sub(r'\s*,\s*', ',', new_val)
                if new_val != val:
                    cell.value = new_val
                    row_issues.append(f'{col_name}: Fixed delimiters')

        # Write all issues in the Comments column
        if row_issues:
            append_issue(row_num, ', '.join(row_issues))

    wb.save(output_path)

if __name__ == '__main__':
    input_file = 'Outdoor-recreation_Scope-Rings-and-Adaptors_reverse_PDW_[by_Sarang-P]_1763041008_ce28de14.xlsx'
    output_file = 'validated_report_output_all_columns9.xlsx'
    run_validations(input_file, output_file)
    print(f'Validation completed. Output saved to {output_file}')


# if __name__ == '__main__':
#     input_file = 'validated_output_with_comments_column.xlsx'  # your input excel file
#     output_file = 'validated_report_output.xlsx'  # output with comments column with issues
#     run_validations(input_file, output_file)
#     print(f'Validation completed. Output saved to {output_file}')
